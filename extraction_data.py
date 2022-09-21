from asyncio.format_helpers import extract_stack
import openpyxl as xl
import pandas
from traceback import print_list
import xlsxwriter


def extract_M(URL_xlsx,nb_frm,nb_plt ):
    wb=xl.load_workbook(URL_xlsx)
    M=[]
    sheet=wb['Sheet'] 
    i=3   
    while i < (nb_frm*2+1)*12+2:
       
        j=0
        MJ=[]
        while j < nb_frm:
            MK = []
            k = 4
            while k-4< nb_plt:
                MK.append(float(sheet.cell(k, i).value))
                k=k+1
                
            MJ.append(MK)
            j=j+1
            i=i+2
        M.append(MJ)
        i=i+1
  
    return M


def extract_P(URL_xlsx, nb_frm, nb_plt):
    wb = xl.load_workbook(URL_xlsx)
    M = []
    sheet = wb['Sheet']
    i = 4
    while i < (nb_frm*2+1)*12+2:

        j = 0
        MJ = []
        while j < nb_frm:
            MK = []
            k = 4
            while k-4 < nb_plt:
                MK.append(float(sheet.cell(k, i).value))
                k = k+1

            MJ.append(MK)
            j = j+1
            i = i+2
        M.append(MJ)
        i = i+1
    return M


def extract_C(URL_xlsx, nb_frm, nb_plt):
    wb = xl.load_workbook(URL_xlsx)
    M = []
    sheet = wb['Sheet']
    i = 3+nb_frm*2
    while i < (nb_frm*2+1)*12+3:
        k=4
        Mk=[]
        while k-4 < nb_plt:
            Mk.append(float(sheet.cell(k, i).value))
            k=k+1
        M.append(Mk)
        i = i+nb_frm*2+1

    return M


def extract_pilote_liste(URL_xlsx, nb_plt):
    wb = xl.load_workbook(URL_xlsx)
    sheet = wb['Sheet']
    k = 4
    LP = []
    while k-4 < nb_plt:
        LP.append(sheet.cell(k, 1).value)
        k = k+1
    
    return LP


def extract_fonction_liste(URL_xlsx, nb_plt):
    wb = xl.load_workbook(URL_xlsx)
    sheet = wb['Sheet']
    k = 4
    LP = []
    while k-4 < nb_plt:
        LP.append(sheet.cell(k, 2).value)
        k = k+1
    return LP

def extract_formation_liste(URL_xlsx, nb_frm):
    wb = xl.load_workbook(URL_xlsx)
    sheet = wb['Sheet']
    k = 3
    LP = []
    while k-3 < nb_frm*2:
        LP.append(sheet.cell(2, k).value)
        k = k+2
    
    return LP




